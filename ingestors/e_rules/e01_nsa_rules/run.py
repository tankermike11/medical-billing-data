"""NSA / GFE-PPDR / Ground-Ambulance Ruleset ingestor (Family E).

Source: counsel-reviewed workbook placed manually at:
  data/raw/e01_nsa_rules/nsa_ruleset.xlsx

This ingestor follows the a07_revenue pattern — no file to download, so it
bypasses run_ingestor() and calls record_release() / record_error() directly.

Expected workbook structure (see Addendum v1.3 §5.2):
  Meta sheet   — key-value rows: ruleset_version, effective_date,
                 last_reviewed, reviewed_by
  Index sheet  — two columns: table name (e.g. "Table 1") → category letter (A–K)
  Rule sheets  — "Table 1" through "Table 10" and "Table K", each with columns:
                 Rule ID, Summary, Prototype logic, System action, Citation,
                 deadline_days, deadline_basis, qpa_dependent, status

Missing columns are handled gracefully:
  status defaults to 'draft', qpa_dependent to 0, deadline fields to None.
This allows the ingestor to run on a partially extended workbook during counsel review.

The ingestor loads status exactly as written in the workbook — it never approves rules.
Re-running against an updated workbook replaces rows cleanly via rule_id (INSERT OR REPLACE).
"""
import re
from pathlib import Path

import openpyxl

from ingestors.core.db import get_conn
from ingestors.core.pipeline import now_utc, record_release, record_error

SOURCE_ID = "e01_nsa_rules"
MANUAL_XLSX = Path(__file__).resolve().parents[3] / "data" / "raw" / "e01_nsa_rules" / "nsa_ruleset.xlsx"

_RULE_SHEET_RE = re.compile(r"^Table\s+(\d+|K)$", re.IGNORECASE)

# Default positional category map if Index sheet is absent.
# Table 1–10 → A–J, Table K → K
_DEFAULT_CATEGORY = {f"Table {i}": chr(ord("A") + i - 1) for i in range(1, 11)}
_DEFAULT_CATEGORY["Table K"] = "K"


def _build_col_map(header_row) -> dict[str, int]:
    return {
        str(cell.value or "").strip().lower(): i
        for i, cell in enumerate(header_row)
    }


def _cell(row, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    v = row[idx].value
    if v is None or str(v).strip().lower() == "none":
        return None
    return v


def _str(row, idx: int | None) -> str | None:
    v = _cell(row, idx)
    return str(v).strip() if v is not None else None


def _int(row, idx: int | None, default: int = 0) -> int:
    v = _cell(row, idx)
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def _read_meta(wb) -> dict[str, str | None]:
    if "Meta" not in wb.sheetnames:
        return {"ruleset_version": None, "effective_date": None,
                "last_reviewed": None, "reviewed_by": None}
    ws = wb["Meta"]
    meta: dict[str, str | None] = {}
    for row in ws.iter_rows(values_only=False):
        if len(row) >= 2 and row[0].value and row[1].value:
            key = str(row[0].value).strip().lower().replace(" ", "_")
            meta[key] = str(row[1].value).strip()
    return {
        "ruleset_version": meta.get("ruleset_version"),
        "effective_date":  meta.get("effective_date"),
        "last_reviewed":   meta.get("last_reviewed"),
        "reviewed_by":     meta.get("reviewed_by"),
    }


def _read_index(wb) -> dict[str, str]:
    if "Index" not in wb.sheetnames:
        print(f"[{SOURCE_ID}] Index sheet absent — using default Table→category mapping")
        return _DEFAULT_CATEGORY.copy()
    ws = wb["Index"]
    index: dict[str, str] = {}
    for row in ws.iter_rows(values_only=False):
        if len(row) >= 2 and row[0].value and row[1].value:
            table_name = str(row[0].value).strip()
            # Index values are like "A. Applicability and routing rules" — extract just the letter
            category_str = str(row[1].value).strip()
            category = category_str[0].upper() if category_str else None
            if category:
                index[table_name] = category
    return index or _DEFAULT_CATEGORY.copy()


def main():
    if not MANUAL_XLSX.exists():
        print(f"[{SOURCE_ID}] MANUAL STEP REQUIRED: place workbook at {MANUAL_XLSX}")
        print("  See module docstring for required workbook structure.")
        return

    conn = get_conn()
    release_id = None

    try:
        wb = openpyxl.load_workbook(MANUAL_XLSX, read_only=True, data_only=True)

        meta  = _read_meta(wb)
        index = _read_index(wb)

        rule_sheets = [name for name in wb.sheetnames if _RULE_SHEET_RE.match(name)]
        print(f"[{SOURCE_ID}] found {len(rule_sheets)} rule sheet(s): {rule_sheets}")

        db_rows = []
        for sheet_name in rule_sheets:
            category = index.get(sheet_name)
            if not category:
                print(f"[{SOURCE_ID}] WARNING: {sheet_name} not in Index — skipping")
                continue

            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=False)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            cm = _build_col_map(header_row)

            rule_id_col       = cm.get("rule id")
            summary_col       = cm.get("summary")
            logic_col         = cm.get("prototype logic")
            action_col        = cm.get("system action")
            citation_col      = cm.get("citation")
            deadline_days_col = cm.get("deadline_days")
            deadline_basis_col= cm.get("deadline_basis")
            qpa_col           = cm.get("qpa_dependent")
            status_col        = cm.get("status")

            for row in rows_iter:
                rule_id = _str(row, rule_id_col)
                if not rule_id:
                    continue

                status = _str(row, status_col) or "draft"
                qpa    = _int(row, qpa_col, default=0)
                if qpa not in (0, 1):
                    qpa = 1 if qpa else 0

                db_rows.append((
                    rule_id,
                    category,
                    _str(row, summary_col),
                    _str(row, logic_col),
                    _str(row, action_col),
                    _str(row, citation_col),
                    _cell(row, deadline_days_col),   # int or None
                    _str(row, deadline_basis_col),
                    qpa,
                    meta["ruleset_version"],
                    meta["effective_date"],
                    meta["last_reviewed"],
                    meta["reviewed_by"],
                    status,
                    SOURCE_ID,
                ))

        wb.close()

        if not db_rows:
            record_error(conn, SOURCE_ID, None, "parse_error",
                         "0 rules loaded — workbook may be empty, lack rule sheets, or use unexpected sheet names")
            print(f"[{SOURCE_ID}] ERROR: 0 rules loaded — see extraction_errors")
            return

        conn.executemany(
            """INSERT OR REPLACE INTO nsa_rules
               (rule_id, category, summary, prototype_logic, system_action, citation,
                deadline_days, deadline_basis, qpa_dependent,
                ruleset_version, effective_date, last_reviewed, reviewed_by,
                status, source_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            db_rows,
        )
        conn.commit()
        release_id = record_release(conn, SOURCE_ID, str(MANUAL_XLSX), "manual_import", len(db_rows))
        print(f"[{SOURCE_ID}] done — {len(db_rows)} rules loaded (release #{release_id})")

    except Exception as e:
        record_error(conn, SOURCE_ID, release_id, "parse_error", str(e))
        raise


if __name__ == "__main__":
    main()
